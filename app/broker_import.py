"""Парсер брокерского отчёта Т-Банка (xlsx).

Отчёт — единый широкий лист с логическими секциями. Каждая «колонка»
занимает несколько ячеек (объединённые), поэтому ориентируемся на тексты
заголовков: находим секцию по названию в колонке A, читаем строку заголовков,
строим карту «индекс колонки → поле» и затем читаем строки данных.

Извлекаем:
- positions — текущие позиции (секция 3.1, «Исходящий остаток»)
- trades    — совершённые сделки (секции 1.1 и 1.3)
"""

import re
from io import BytesIO

import openpyxl

# Заголовок секции: «1.1 Текст», «2. Текст», «3.1 Текст» — номер, пробел, затем не-цифра.
# Дата «27.11.2024» не подходит (нет пробела с текстом после номера).
_SECTION_RE = re.compile(r"^\d+(\.\d+)?\.?\s+\D")


def _norm(value) -> str:
    """Нормализует текст заголовка: убирает пробелы/переводы строк, в нижний регистр."""
    if value is None:
        return ""
    return "".join(str(value).split()).lower()


def _find_section_row(ws, title_substr: str) -> int | None:
    target = title_substr.lower()
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a and target in str(a).lower():
            return r
    return None


def _build_column_map(ws, header_row: int, targets: dict[str, set[str]]) -> dict[str, int]:
    """targets: поле → варианты нормализованного заголовка. Возвращает поле → индекс колонки."""
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        norm = _norm(ws.cell(row=header_row, column=c).value)
        if not norm:
            continue
        for field, variants in targets.items():
            if field not in col_map and norm in variants:
                col_map[field] = c
    return col_map


def _is_section_title(value) -> bool:
    """Строка вида «2. ...», «3.1 ...» — заголовок секции (конец данных)."""
    if not value:
        return False
    return bool(_SECTION_RE.match(str(value).strip()))


def _read_rows(ws, header_row: int, col_map: dict[str, int], required_field: str):
    """Читает строки данных после заголовка до следующей секции. Пустые/служебные пропускаются."""
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if _is_section_title(a):
            break
        record = {}
        for field, col in col_map.items():
            record[field] = ws.cell(row=r, column=col).value
        if record.get(required_field) in (None, ""):
            continue  # служебная строка (разрыв страницы) или пустая
        rows.append(record)
    return rows


_TRADE_TARGETS = {
    "date": {"датазаключения"},
    "time": {"время"},
    "side": {"видсделки"},
    "name": {"наименованиеактива"},
    "ticker": {"кодактива"},
    "price": {"ценазаединицу"},
    "price_currency": {"валютацены"},
    "quantity": {"количество"},
    "amount": {"суммасделки"},
    "settle_currency": {"валютарасчетов"},
    "commission": {"комиссияброкера"},
}

_POSITION_TARGETS = {
    "name": {"наименованиеактива"},
    "ticker": {"кодактива"},
    "isin": {"isin"},
    "quantity": {"исходящийостаток"},
}


def _to_float(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _clean_ticker(v) -> str:
    """Нормализует код актива: убирает суффикс борда после «@» (напр. TRUR@ → TRUR)."""
    return str(v or "").strip().split("@")[0].strip()


# Валютные/металлические пары (конвертация валют, а не покупка бумаг) — в P&L не учитываются
_FX_RE = re.compile(r"(_TOM|_TDM|_TOD)$|^(USD|EUR|CNY|GBP|CHF|HKD|JPY|GLD|SLV)\w*RUB|USD000")


def is_fx_ticker(ticker: str) -> bool:
    return bool(_FX_RE.search(ticker or ""))


def _parse_date_any(s: str):
    """Парсит дату из строки DD.MM.YYYY → сравнимый ключ (или пустая строка)."""
    s = str(s or "").strip()
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else ""


_KNOWN_CURRENCIES = {"RUB", "USD", "EUR", "CNY", "GBP", "CHF", "JPY", "HKD", "KZT", "TRY"}


def _parse_cash_balances(ws) -> dict[str, float]:
    """Свободные денежные остатки по валютам (таблица балансов секции 2, «Исходящий остаток»)."""
    start = _find_section_row(ws, "операции с денежными")
    if start is None:
        return {}
    out: dict[str, float] = {}
    header = None
    for r in range(start + 1, min(start + 20, ws.max_row + 1)):
        a = ws.cell(row=r, column=1).value
        if a and "Валюта" in str(a):
            header = r
            break
    if header is None:
        return {}
    for r in range(header + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        code = str(a).strip() if a else ""
        closing = ws.cell(row=r, column=21).value  # колонка U — «Исходящий остаток»
        if code not in _KNOWN_CURRENCIES or closing is None:
            break  # конец таблицы балансов
        amt = _to_float(closing)
        if amt:
            out[code] = amt
    return out


def _parse_cashflows(ws) -> list[dict]:
    """Дивиденды/купоны и налоги из секции «2. Операции с денежными средствами»."""
    start = _find_section_row(ws, "операции с денежными")
    if start is None:
        return []
    out = []
    for r in range(start + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a and _is_section_title(a) and not str(a).strip().startswith("2"):
            break
        op = ws.cell(row=r, column=38).value  # колонка AL — «Операция»
        if not op or str(op).strip() == "Операция":
            continue
        op_l = str(op).strip().lower()
        date = _parse_date_any(ws.cell(row=r, column=23).value or ws.cell(row=r, column=1).value)
        credit = _to_float(ws.cell(row=r, column=53).value)  # BA — зачисление
        debit = _to_float(ws.cell(row=r, column=66).value)  # BN — списание
        if "выплата доходов" in op_l and credit:
            # Начисления (дивиденды/купоны) — зачислены уже за вычетом налога с дивидендов
            out.append({"date": date, "kind": "dividend", "amount": round(credit, 2)})
        elif "комиссия" in op_l and debit:
            out.append({"date": date, "kind": "commission", "amount": round(debit, 2)})
        elif "налог" in op_l:
            # Налог с дивидендов уже отражён в чистых начислениях — в строку «налоги» не идёт
            if "дивиденд" not in op_l:
                out.append({"date": date, "kind": "tax", "amount": round(credit - debit, 2)})
        elif "пополнение счета" in op_l and credit:
            out.append({"date": date, "kind": "deposit", "amount": round(credit, 2)})
        elif "вывод средств" in op_l and debit:
            out.append({"date": date, "kind": "withdrawal", "amount": round(debit, 2)})
    return out


def _parse_trades_section(ws, title: str) -> list[dict]:
    start = _find_section_row(ws, title)
    if start is None:
        return []
    header_row = start + 1
    col_map = _build_column_map(ws, header_row, _TRADE_TARGETS)
    if "ticker" not in col_map:
        return []
    out = []
    for rec in _read_rows(ws, header_row, col_map, "ticker"):
        out.append(
            {
                "date": str(rec.get("date") or "").strip(),
                "time": str(rec.get("time") or "").strip(),
                "side": str(rec.get("side") or "").strip(),
                "name": str(rec.get("name") or "").strip(),
                "ticker": _clean_ticker(rec.get("ticker")),
                "price": _to_float(rec.get("price")),
                "currency": str(
                    rec.get("settle_currency") or rec.get("price_currency") or ""
                ).strip(),
                "quantity": _to_float(rec.get("quantity")),
                "amount": _to_float(rec.get("amount")),
                "commission": _to_float(rec.get("commission")),
            }
        )
    return out


def parse_broker_report(file_bytes: bytes) -> dict:
    """Парсит xlsx-отчёт Т-Банка. Возвращает {positions, trades, cashflows, report_date}."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # --- Позиции (3.1 Движение по ценным бумагам) ---
    positions = []
    pos_start = _find_section_row(ws, "движение по ценным бумаг")
    if pos_start is not None:
        header_row = pos_start + 1
        col_map = _build_column_map(ws, header_row, _POSITION_TARGETS)
        if "ticker" in col_map:
            for rec in _read_rows(ws, header_row, col_map, "ticker"):
                qty = _to_float(rec.get("quantity"))
                if qty <= 0:
                    continue  # позиция закрыта
                positions.append(
                    {
                        "name": str(rec.get("name") or "").strip(),
                        "ticker": _clean_ticker(rec.get("ticker")),
                        "isin": str(rec.get("isin") or "").strip(),
                        "quantity": qty,
                    }
                )

    # --- Сделки (1.1 совершённые + 1.3 за расчётный период), дедуп по ключу ---
    trades = []
    seen = set()
    report_date = ""
    for title in ("информация о совершенных", "сделки за расчетный"):
        for t in _parse_trades_section(ws, title):
            key = (t["date"], t["time"], t["ticker"], t["side"], t["quantity"], t["amount"])
            if key in seen:
                continue
            seen.add(key)
            t["is_fx"] = is_fx_ticker(t["ticker"])
            trades.append(t)
            d = _parse_date_any(t["date"])
            if d > report_date:
                report_date = d

    cashflows = _parse_cashflows(ws)
    for cf in cashflows:
        if cf["date"] > report_date:
            report_date = cf["date"]

    cash = _parse_cash_balances(ws)

    wb.close()
    return {
        "positions": positions,
        "trades": trades,
        "cashflows": cashflows,
        "cash": cash,
        "report_date": report_date,
    }
