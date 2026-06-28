"""Импорт брокерского отчёта: парсер и эндпоинты."""

from io import BytesIO

import openpyxl

from app.broker_import import parse_broker_report


def _build_sample_report() -> bytes:
    """Минимальный xlsx, повторяющий структуру отчёта Т-Банка (секции 1.1 и 3.1)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # --- Секция 1.1: совершённые сделки ---
    ws["A1"] = "1.1 Информация о совершенных сделках"
    headers = [
        "Дата заключения",
        "Время",
        "Вид сделки",
        "Наименование актива",
        "Код актива",
        "Цена за единицу",
        "Валюта цены",
        "Количество",
        "Сумма сделки",
        "Валюта расчетов",
        "Комиссия брокера",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    trade = [
        "27.11.2024",
        "12:24",
        "Покупка",
        "ГАЗПРОМ ао",
        "GAZP",
        114.57,
        "RUB",
        10,
        1145.7,
        "RUB",
        3.44,
    ]
    for i, v in enumerate(trade, start=1):
        ws.cell(row=3, column=i, value=v)

    # --- Секция 3.1: движение по ценным бумагам ---
    ws["A4"] = "3.1 Движение по ценным бумагам"
    pos_headers = ["Наименование актива", "Код актива", "ISIN", "Исходящий остаток"]
    for i, h in enumerate(pos_headers, start=1):
        ws.cell(row=5, column=i, value=h)
    ws.cell(row=6, column=1, value="ГАЗПРОМ ао")
    ws.cell(row=6, column=2, value="GAZP")
    ws.cell(row=6, column=3, value="RU0007661625")
    ws.cell(row=6, column=4, value=10)
    # закрытая позиция (остаток 0) — должна отфильтроваться
    ws.cell(row=7, column=1, value="Закрытая бумага")
    ws.cell(row=7, column=2, value="CLSD")
    ws.cell(row=7, column=3, value="RU000CLOSED0")
    ws.cell(row=7, column=4, value=0)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parser_extracts_positions_and_trades():
    res = parse_broker_report(_build_sample_report())
    assert len(res["positions"]) == 1  # закрытая позиция отфильтрована
    pos = res["positions"][0]
    assert pos["ticker"] == "GAZP"
    assert pos["quantity"] == 10
    assert pos["isin"] == "RU0007661625"

    assert len(res["trades"]) == 1
    t = res["trades"][0]
    assert t["side"] == "Покупка"
    assert t["ticker"] == "GAZP"
    assert t["price"] == 114.57
    assert t["quantity"] == 10


def _make_portfolio(client, token, name="Брокер") -> int:
    """Создаёт портфель и возвращает его id."""
    resp = client.post(
        "/api/portfolios",
        params={"token": token},
        json={"name": name},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()["id"]


def _xlsx_files(report):
    return {
        "files": (
            "report.xlsx",
            report,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


def test_reimport_dedupes_trades(client, registered):
    """Повторная загрузка того же отчёта не создаёт дублей сделок (слияние)."""
    token = registered["access_token"]
    pid = _make_portfolio(client, token)
    report = _build_sample_report()
    params = {"token": token, "portfolio_id": pid}
    first = client.post("/api/portfolio/import", params=params, files=_xlsx_files(report))
    assert first.json()["new_trades"] == 1
    second = client.post("/api/portfolio/import", params=params, files=_xlsx_files(report))
    assert second.json()["new_trades"] == 0  # дубли не добавились

    got = client.get("/api/portfolio", params={"token": token, "portfolio_id": pid}).json()
    assert len(got["trades"]) == 1


def test_import_and_get_portfolio(client, registered):
    token = registered["access_token"]
    pid = _make_portfolio(client, token)
    report = _build_sample_report()

    resp = client.post(
        "/api/portfolio/import",
        params={"token": token, "portfolio_id": pid},
        files=_xlsx_files(report),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["new_trades"] == 1

    got = client.get("/api/portfolio", params={"token": token, "portfolio_id": pid})
    assert got.status_code == 200
    data = got.json()
    assert data["has_data"] is True
    assert data["positions"][0]["ticker"] == "GAZP"
    assert data["trades"][0]["side"] == "Покупка"


def test_aggregate_sums_portfolios(client, registered):
    """portfolio_id=all суммирует позиции одинаковых тикеров из разных портфелей."""
    token = registered["access_token"]
    p1 = _make_portfolio(client, token, "Первый")
    p2 = _make_portfolio(client, token, "Второй")
    report = _build_sample_report()
    client.post(
        "/api/portfolio/import",
        params={"token": token, "portfolio_id": p1},
        files=_xlsx_files(report),
    )
    client.post(
        "/api/portfolio/import",
        params={"token": token, "portfolio_id": p2},
        files=_xlsx_files(report),
    )
    agg = client.get("/api/portfolio", params={"token": token, "portfolio_id": "all"}).json()
    gazp = next(p for p in agg["positions"] if p["ticker"] == "GAZP")
    assert gazp["quantity"] == 20  # 10 + 10


def test_aggregate_ignores_orphan_rows(client, registered):
    """«Общий капитал» не учитывает строки без portfolio_id (старые импорты)."""
    from app.db import get_db_connection

    token = registered["access_token"]
    email = registered["email"]
    pid = _make_portfolio(client, token)
    report = _build_sample_report()
    client.post(
        "/api/portfolio/import",
        params={"token": token, "portfolio_id": pid},
        files=_xlsx_files(report),
    )
    # Осиротевшая позиция тем же тикером (как от импорта до мультипортфелей)
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO portfolio_positions (email, portfolio_id, ticker, name, isin, quantity) "
        "VALUES (%s, NULL, 'GAZP', 'ГАЗПРОМ ао', 'RU0007661625', 99)",
        (email,),
    )
    conn.commit()
    conn.close()

    agg = client.get("/api/portfolio", params={"token": token, "portfolio_id": "all"}).json()
    gazp = next(p for p in agg["positions"] if p["ticker"] == "GAZP")
    assert gazp["quantity"] == 10  # только привязанная позиция, осиротевшая (99) исключена


def test_import_preview_reports_summary_without_saving(client, registered):
    """Предпросмотр возвращает сводку и не пишет в портфель."""
    token = registered["access_token"]
    pid = _make_portfolio(client, token)
    report = _build_sample_report()
    prev = client.post(
        "/api/portfolio/import/preview",
        params={"token": token, "portfolio_id": pid},
        files=_xlsx_files(report),
    )
    assert prev.status_code == 200, prev.text
    s = prev.json()
    assert s["new_trades"] == 1
    assert s["assets"] == 1
    assert s["period"]["from"] == "27.11.2024"
    # ничего не сохранилось
    got = client.get("/api/portfolio", params={"token": token, "portfolio_id": pid}).json()
    assert got["has_data"] is False


def test_import_accepts_multiple_files(client, registered):
    """Можно загрузить несколько файлов разом; дубли между ними схлопываются."""
    token = registered["access_token"]
    pid = _make_portfolio(client, token)
    report = _build_sample_report()
    files = [
        (
            "files",
            ("a.xlsx", report, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ),
        (
            "files",
            ("b.xlsx", report, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ),
    ]
    resp = client.post(
        "/api/portfolio/import",
        params={"token": token, "portfolio_id": pid},
        files=files,
    )
    assert resp.status_code == 200, resp.text
    # одна и та же сделка в двух файлах → одна новая
    assert resp.json()["new_trades"] == 1
    got = client.get("/api/portfolio", params={"token": token, "portfolio_id": pid}).json()
    assert len(got["trades"]) == 1


def test_calendar_endpoint_shape(client, registered):
    """Эндпоинт календаря отвечает корректной структурой (в тестах сеть выключена)."""
    token = registered["access_token"]
    pid = _make_portfolio(client, token)
    report = _build_sample_report()
    client.post(
        "/api/portfolio/import",
        params={"token": token, "portfolio_id": pid},
        files=_xlsx_files(report),
    )
    r = client.get("/api/portfolio/calendar", params={"token": token, "portfolio_id": pid})
    assert r.status_code == 200, r.text
    data = r.json()
    assert set(data) >= {"events", "by_month", "total", "currency"}
    assert isinstance(data["events"], list)
    assert data["total"] == 0.0  # QUOTES_ENABLED=false → выплаты не запрашиваются


def test_import_rejects_non_xlsx(client, registered):
    token = registered["access_token"]
    pid = _make_portfolio(client, token)
    resp = client.post(
        "/api/portfolio/import",
        params={"token": token, "portfolio_id": pid},
        files={"files": ("report.txt", b"not a spreadsheet", "text/plain")},
    )
    assert resp.status_code == 400


def test_portfolio_requires_valid_token(client):
    resp = client.get("/api/portfolio", params={"token": "garbage.token"})
    assert resp.status_code == 401
